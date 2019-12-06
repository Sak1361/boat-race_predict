import string as stg    #アルファベット用
import requests, re, sys, os, time
from bs4 import BeautifulSoup   #scrape
import openpyxl as opxl     #Excel
from openpyxl.styles.borders import Border, Side    #Excelの線引き
from openpyxl.styles.alignment import Alignment #中央揃え

class Scraping_page:
    def __init__(self,place,day,rno):
        self.page = rno
        self.date = day
        self.dir_name = f'fukuoka_{day}'
        self.scrape_cls = ['racer_name','boat_result','boat_course',
        'course_result','course_winning_tech','pool_result']
        self.search_place = {'桐生':'01','戸田':'02','江戸川':'03','平和島':'04',
        '多摩川':'05','浜名湖':'06','蒲郡':'07','常滑':'08','津':'09','三国':'10',
        'びわこ':'11','住之江':'12','尼崎':'13','鳴門':'14','丸亀':'15','児島':'16',
        '宮島':'17','徳山':'18','下関':'19','若松':'20','芦屋':'21','福岡':'22',
        '唐津':'23','大村':'24'}
        self.selects = self.search_place[place]

    def crawling(self):
        pages = self.page
        #レース情報URL
        url = f'https://www.boatrace.jp/owpc/pc/race/racelist?rno={pages}&jcd={self.selects}&hd={self.date}'
        #ディレクトリ作成部
        race_num = f'{pages}R'
        if not os.path.exists(f'{self.dir_name}/{race_num}'):
            os.mkdir(f'{self.dir_name}/{race_num}')
        f_name = f'{self.dir_name}/{race_num}/{race_num}.html'
        #ページ保存部
        self.craw_page(url,f_name)
        ##個人成績取得
        with open(f_name,'r') as html:
            soup = BeautifulSoup(html, "html.parser")
            racer_n = 1
            for boat_number in soup.select('div[class="is-fs11"]'):     #divのis-fs11完全一致のみセレクト
                boat_number = re.sub('[ /]','',boat_number.text)    #textにして空白と/を除外
                boat_number = boat_number.split('\n')   #\nで区切ってリスト化
                for b_n in boat_number:     #改行＋番号＋階級と出るから番号のみ取得
                    try:
                        b_num = int(b_n)
                        nothing = 0
                        break
                    except ValueError:
                        pass
                    nothing = 1 #ここまできた==無い
                if nothing:     #同タグで出身＋年齢もあるのでその場合はスキップ
                    continue
                urls = [ f"http://www.boatrace-db.net/racer/yresult/regno/{b_num}/year/2019/",  #今期選手情報（期別から検索
                    f'http://www.boatrace-db.net/racer/aresult/regno/{b_num}/',   #通算成績
                    f'http://www.boatrace-db.net/racer/rdemo/regno/{b_num}/']    #スタート順位と成績
                d_n = ['2019','all','pre-times']
                for i in range(len(urls)):
                    f_racer = f'{self.dir_name}/{race_num}/boat{racer_n}_{d_n[i]}.html'
                    self.craw_page(urls[i],f_racer)   #クローリング
                    time.sleep(2) #時間を開けないとダメだってさ
                print(f'Get {race_num}-{racer_n}st')
                racer_n += 1
            if racer_n == 1:    #レーサー取得できなかった場合
                print('Page not found. exit to program.')
                sys.exit()
            else:
                print(f'Get {pages}R')
            #ページ遷移
            pages += 1
            if pages > 12:
                return None
            self.page = pages

        self.crawling()

    def craw_page(self,url,f_name):
        headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) \
        AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36'}
        response = requests.get(url,headers=headers)
        response.encoding = response.apparent_encoding
        with open(f_name,'w',encoding="utf-8") as f:
            f.write(response.text)

    def scrap_racer(self,race):
        res = []
        tag_select = ['odd','even']    #奇数タグと偶数タグ
        round_lambda = lambda x:int((x*2+1)//2) #偶数丸め込みだと0.5を0にするため
        meet = round_lambda(int(self.selects)/2)    #何番目か
        d_n = ['2019','pre-times','all']
        if int(self.selects) % 2:
            tag = tag_select[0] #割り切れないので奇数
        else:
            tag = tag_select[1] #割り切れるので偶数
        for racer in range(1,7):
            for j in range(len(d_n)):
                dir_p = f"{self.dir_name}/{race}R/boat{racer}_{d_n[j]}.html"
                with open(dir_p,'r') as html:
                    soup = BeautifulSoup(html, "html.parser")
                    if d_n[j]=='pre-times':    #'pre-times'のみ異なる
                        time_res = soup.find(class_='rdemo').text
                        res.append(self.shaping(time_res))
                    else:
                        for cnt in range(6):    #名前と成績を取得
                            if cnt==5:  #福岡勝率だけ抜き出す
                                j = 1
                                tmp_cls = soup.find(class_=self.scrape_cls[cnt])
                                strings = tmp_cls.find(class_='header').text
                                for txt in tmp_cls.find_all(class_=tag):   #タグが奇数偶数で分かれてる
                                    if j==meet:
                                        txt = txt.text
                                        strings = strings + txt
                                        res.append(self.shaping(strings))
                                        j += 1
                                    else:
                                        j += 1
                            else:   #クラス名ずつにスクレイプ
                                tmp_cls = soup.find(class_=self.scrape_cls[cnt])
                                if tmp_cls == None:
                                    strings = soup.find(class_='side_title').text  #racer_nameクラスがない場合
                                    if strings == None:
                                        strings = f'{cnt}号艇'
                                else:
                                    strings = soup.find(class_=self.scrape_cls[cnt]).text  #タグが入るのでtext
                                res.append(self.shaping(strings))

        return res

    def shaping(self,strings):
        tmp_res = []
        strings = strings.replace(' ','') #空白除去
        strings = strings.split('\n')
        for i in strings:    #空文字除去
            if i != '':
                tmp_res.append(i)
        return tmp_res


class Excel:
    def __init__(self):
        self.row = 2   #縦
        self.column = 2  #横
        self.column_init = 2  #横初期値
        self.capa = 0   #横に２つずつ書くため
        self.width = 0
        self.title = ['【通算成績】','【2019成績】']
        self.count = 0

    def write_xl(self,data,xls):
        if len(data) == 1:  #レーサー氏名入力
            if self.count:
                self.row += 8

            self.column = self.column_init
            self.capa = 0
            data = re.sub(r'[!-~\s+]','',data[0]) #空白文字が混ざってる時があるので除く
            xls.cell(row=self.row,column=1,value=data) #書き込み
            if not self.count:
                xls.cell(row=self.row+1,column=1,value=self.title[1]) #見出し2019
            else:
                xls.cell(row=self.row,column=1,value=self.title[0]) #見出し通算
            self.count=0 if self.count else 1   #0-1入れ替え

        else:
            size = 4   #表を何個横に置くか
            #幅指定 基本7行(6艇+項目名)+A列あけ
            if len(data)/8 == 7:    #wid==8,leng==7
                width = 9
            elif len(data)/9 == 7:  #wid==9,leng==7
                width = 10
            else:   #その他
                width = 12

            if self.column-2 == 12: #前回がwidth==12の場合
                self.column += 7
            if self.capa:
                width += self.column-2

            row = self.row
            column = self.column

            for d in data:
                try:
                    d = int(d)  #なるだけ数字は数値に
                except ValueError:
                    try:
                        d = float(d)    #小数値はfloatで変換
                    except ValueError:
                        pass               
                xls.cell(row=row,column=column,value=d) #書き込み
                if column==width:
                    row+=1
                    column = self.column
                else:
                    column+=1
            if self.capa == size-1:
                self.row += 8
                self.column = self.column_init
                self.capa = 0
            else:
                self.column = width + 2
                self.capa += 1
            ###else_end###
        return xls

    def add_frame(self,all_sheet):
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
        pass_A = re.compile(r'A[0-9]')   #A列のみ省きたいので
        for sheet in all_sheet: #シートごとに枠線付け
            for row in sheet:
                for cell in row:    #セルごとに回す
                    if sheet[cell.coordinate].value != None:    #該当のセルがNone以外==何かしらある
                        cell.alignment = Alignment(horizontal = 'center',   #全てを中央揃えに
                                    vertical = 'center',
                                    wrap_text = False)
                        if pass_A.match(cell.coordinate):  #A行には枠線をつけない
                            continue
                        sheet[cell.coordinate].border = border  #枠線つけ

    def adjust_width(self,all_sheet):
        alph = list(stg.ascii_uppercase)
        for sheet in all_sheet: #シートごとに
            for col in sheet.columns:
                max_length = 0
                column = col[0].column
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjust = (max_length + 2) * 1.2
                if column > 26:
                    calc = int(column/27)
                    col_alph = f'{alph[calc]}{alph[column-(calc*column)-1]}'
                else:
                    col_alph = f'{alph[column-1]}'
                sheet.column_dimensions[col_alph].width = adjust

if __name__ == "__main__":
    place = sys.argv[1]
    day = sys.argv[2]
    race_number_start = int(sys.argv[3])
    race_number_end = int(sys.argv[4])
    exist = False
    if not os.path.exists('fukuoka_'+day):  #dir作成
        os.mkdir('fukuoka_'+day)
    else:
        exist = True
    scrape = Scraping_page(place,int(day),race_number_start)    #引数に場所,日,レース番号(8なら8レースから)
    if exist:
        cho = int(input('directry is exist.crawling again?(0 or 1):') )
        if cho:
            print("Interval of 2.5 seconds per page due to website restrictions.")
            scrape.crawling()    #クローリング
        else:
            print('skip crawling.')
    else:
        print("New acquisition\nInterval of 2.5 seconds per page due to website restrictions.")
        scrape.crawling()    #クローリング

    print('Start scraping and write Excel')
    xl = opxl.Workbook()   #新規作成
    for _i in range(race_number_start,race_number_end+1):
        xl.create_sheet(title=f'{_i}R')
    xl.remove(xl.worksheets[0]) #空のsheet1を削除(元からつくりたくないけど)
    for rn in range(race_number_start,race_number_end+1):
        excel = Excel() #ページ毎に初期化
        xs = xl.active
        xs = xl[f'{rn}R']    #シート指定
        #xs.title = f'{rn}R'  #シート名変更
        xs['A1'].value = f'【福岡 {rn}R】'
        for data in scrape.scrap_racer(rn):
            xs = excel.write_xl(data,xs)
        print(f"Page-{rn} completed.")
    
    excel.add_frame(xl.worksheets)  #表の枠線つけ部
    excel.adjust_width(xl.worksheets)

    xl.save(f"/Users/sak1361/repository/boat_race/{day}.xlsx")   #保存
    print(f'.xlsx file saved~\n/Users/sak1361/repository/boat_race/{day}.xlsx')

    #ws = wb.get_sheet_by_name('SHEETNAME') #SHEETNAMEを検索して指定

    #http://www.boatrace-db.net/stadium/motor/pid/22/   モータ勝率  下記のトップ
    #http://www.boatrace-db.net/stadium/boat/pid/22/    ボート
    #http://www.boatrace-db.net/stadium/tcourse/pid/22/     コース別総合
    #http://www.boatrace-db.net/stadium/ccourse/pid/22/     条件別（風やSG
    #http://www.boatrace-db.net/stadium/result/pid/22/      出目
    #http://www.boatrace-db.net/stadium/demo/pid/22/        展示タイム別
#ボートレース公式サイト
    #https://www.boatrace.jp/owpc/pc/extra/data/download.html   公式の選手成績
    #https://www.boatrace.jp/owpc/pc/data/stadium?jcd=22&hd=20190920    場データ jcd=開催状,hd=年日
    # https://www.boatrace.jp/owpc/pc/race/index?hd=20190916    レース一覧
    #https://www.boatrace-fukuoka.com/modules/datafile/?page=index_mrankdtl&kind=1&type=1   福岡モータ成績
    #https://www.boatrace-fukuoka.com/modules/datafile/?page=index_tanpyou  選手別戦評（多分当節のみ
    #https://www.boatrace.jp/owpc/pc/race/pcexpect?rno=1&jcd=22&hd=20190916     予想

    '''
    レース場	場コード
    桐生	01
    戸田	02
    江戸川	03
    平和島	04
    多摩川	05
    浜名湖	06
    蒲郡	07
    常滑	08
    津	09
    三国	10
    びわこ	11
    住之江	12
    尼崎	13
    鳴門	14
    丸亀	15
    児島	16
    宮島	17
    徳山	18
    下関	19
    若松	20
    芦屋	21
    福岡	22
    唐津	23
    大村	24
    '''
